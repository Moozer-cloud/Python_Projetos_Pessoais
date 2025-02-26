import sqlite3
from abc import ABC, abstractmethod
from datetime import datetime, timedelta
import os
import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import Menu, messagebox
from tkinter import ttk, messagebox

# Conexão com o banco de dados
def conectar_db():
    conn = sqlite3.connect('oficina.db')
    return conn

# Criação das tabelas no banco de dados
def criar_tabelas():
    conn = conectar_db()
    cursor = conn.cursor()
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS clientes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        cpf TEXT NOT NULL UNIQUE,
        endereco TEXT NOT NULL
    )
    ''')
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS veiculos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente_id INTEGER,
        modelo TEXT NOT NULL,
        placa TEXT NOT NULL UNIQUE,
        fabricante TEXT,
                   -- Certifique-se de que a coluna fabricante está aqui
        ano INT,
        FOREIGN KEY (cliente_id) REFERENCES clientes (id)
    )
    ''')
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS servicos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        veiculo_id INTEGER,
        descricao TEXT NOT NULL,
        preco REAL NOT NULL,
        data_servico TEXT NOT NULL,
        data_garantia TEXT NOT NULL,
        FOREIGN KEY (veiculo_id) REFERENCES veiculos (id)
    )
    ''')
    
    conn.commit()
    conn.close()

class Cliente:
    def __init__(self, nome, cpf, endereco):
        self.nome = nome
        self.cpf = cpf
        self.endereco = endereco
        self.id = None  # Inicializa o ID como None

    def salvar(self):
        conn = conectar_db()
        cursor = conn.cursor()
        cursor.execute('INSERT INTO clientes (nome, cpf, endereco) VALUES (?, ?, ?)', 
                       (self.nome, self.cpf, self.endereco))
        conn.commit()
        self.id = cursor.lastrowid  # Recupera o ID do último cliente inserido
        conn.close()

class Veiculo:
    def __init__(self, cliente_id, modelo, placa, fabricante, ano):
        self.cliente_id = cliente_id
        self.modelo = modelo
        self.placa = placa
        self.fabricante = fabricante
        self.ano = ano

    def salvar(self):
        conn = conectar_db()
        cursor = conn.cursor()
        cursor.execute('INSERT INTO veiculos (cliente_id, modelo, placa, fabricante, ano) VALUES (?, ?, ?, ?, ?)',
                       (self.cliente_id, self.modelo, self.placa, self.fabricante, self.ano))
        conn.commit()
        conn.close()

class Servico:
    def __init__(self, veiculo_id, descricao, preco, data_servico, data_garantia):
        self.veiculo_id = veiculo_id
        self.descricao = descricao
        self.preco = preco
        self.data_servico = data_servico
        self.data_garantia = data_garantia  # Nova data de garantia

    def salvar(self):
        conn = conectar_db()
        cursor = conn.cursor()
        cursor.execute('INSERT INTO servicos (veiculo_id, descricao, preco, data_servico, data_garantia) VALUES (?, ?, ?, ?, ?)', 
                       (self.veiculo_id, self.descricao, self.preco, self.data_servico, self.data_garantia))
        conn.commit()
        conn.close()

def menu():
    return input("""\n
    ================ MENU OFICINA ================
    [1]\tCadastrar Cliente
    [2]\tCadastrar Veículo
    [3]\tRegistrar Serviço
    [4]\tListar Clientes
    [5]\tListar Veículos
    [6]\tListar Serviços
    [7]\tExcluir Cliente
    [8]\tExcluir Veículo
    [9]\tRelatórios
    [q]\tSair
    => """)

def cadastrar_cliente():
    nome = input("Informe o nome do cliente: ")
    cpf = input("Informe o CPF do cliente: ")
    endereco = input("Informe o endereço do cliente: ")

    # Verificar se o CPF já está cadastrado
    conn = conectar_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM clientes WHERE cpf = ?', (cpf,))
    cliente_existente = cursor.fetchone()
    conn.close()

    if cliente_existente:
        print("@@@ Já existe um cliente cadastrado com esse CPF! @@@")
        return

    cliente = Cliente(nome, cpf, endereco)
    cliente.salvar()
    print(f"Cliente cadastrado com sucesso! Cliente cadastrado com ID: {cliente.id}")

def cadastrar_veiculo():
    cliente_id = int(input("Informe o ID do cliente: "))
    modelo = input("Informe o modelo do veículo: ")
    placa = input("Informe a placa do veículo: ")
    fabricante = input("Informe o fabricante do veículo: ")
    ano = int(input("Informe o ano do veículo: "))  # Certifique-se de que o ano é um número inteiro
    # Cria uma nova instância da classe Veiculo
    veiculo = Veiculo(cliente_id, modelo, placa, fabricante, ano)
    veiculo.salvar()  # Chama o método para salvar no banco de dados
    print("Veículo cadastrado com sucesso!")

def registrar_servico():
    veiculo_id = int(input("Informe o ID do veículo: "))
    descricao = input("Informe a descrição do serviço: ")
    preco = float(input("Informe o preço do serviço: "))    
    # Solicita a data no formato DD/MM/YYYY
    data_servico = input("Informe a data do serviço (DD/MM/YYYY): ")    
    # Converte a data para o formato YYYY-MM-DD
    data_servico_formatada = datetime.strptime(data_servico, "%d/%m/%Y").strftime("%Y-%m-%d")    
    # Solicita a garantia em dias
    dias_garantia = int(input("Informe a garantia do serviço em dias: "))    
    # Calcula a data de expiração da garantia
    data_garantia = datetime.strptime(data_servico_formatada, "%Y-%m-%d") + timedelta(days=dias_garantia)
    data_garantia_formatada = data_garantia.strftime("%Y-%m-%d")  # Formato para o banco de dados
    
    servico = Servico(veiculo_id, descricao, preco, data_servico_formatada, data_garantia_formatada)
    servico.salvar()
    print("Serviço cadastrado com sucesso!")

def listar_clientes():
    conn = conectar_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM clientes')
    clientes = cursor.fetchall()
    conn.close()

    # Mensagem informativa
    print("\nClientes Cadastrados no Sistema")
    print("=" * 100)
    # Exibe os resultados em formato de tabela
    print(f"{'ID':^5} {'Nome':^30} {'CPF':^15} {'Endereço':^30}")
    print("=" * 100)    
    for cliente in clientes:
        id_cliente, nome, cpf, endereco = cliente
        print(f"{id_cliente:^5} {nome:^30} {cpf:^15} {endereco:^30}")

def listar_veiculos():
    conn = conectar_db()
    cursor = conn.cursor()
    
    # Faz uma junção entre as tabelas veiculos e clientes
    cursor.execute('''
    SELECT v.id, v.modelo, v.placa, v.fabricante, v.ano, c.id AS cliente_id, c.nome
    FROM veiculos v
    JOIN clientes c ON v.cliente_id = c.id
    ''')
    
    veiculos = cursor.fetchall()
    conn.close()

    # Mensagem informativa
    print("\nVeículos Cadastrados no Sistema")
    print("=" * 115)
    # Exibe os resultados em formato de tabela
    print(f"{'ID Veículo':^15} {'Modelo':^20} {'Placa':^10} {'Fabricante':^15} {'Ano':^5} {'ID Cliente':^15} {'Nome do Cliente':^30}")
    print("=" * 115)    
    for veiculo in veiculos:
        id_veiculo, modelo, placa, fabricante, ano, cliente_id, nome_cliente = veiculo
        print(f"{id_veiculo:^15} {modelo:^20} {placa:^10} {fabricante:^15} {ano:^5} {cliente_id:^15} {nome_cliente:^30}")

def listar_servicos():
    conn = conectar_db()
    cursor = conn.cursor()
    
    # Faz uma junção entre as tabelas servicos, veiculos e clientes
    cursor.execute('''
    SELECT s.id, s.descricao, s.preco, s.data_servico, s.data_garantia, v.id AS veiculo_id, v.modelo, c.nome
    FROM servicos s
    JOIN veiculos v ON s.veiculo_id = v.id
    JOIN clientes c ON v.cliente_id = c.id
    ''')
    
    servicos = cursor.fetchall()
    conn.close()

    # Mensagem informativa
    print("\nServiços Cadastrados no Sistema")
    print("=" * 160)
    
    # Exibe os resultados em formato de tabela
    print(f"{'ID Serviço':^15} {'Descrição do Serviço':^30} {'Data do Serviço':^15} {'Data Garantia':^15} {'Preço':^10} {'ID Veículo':^15} {'Modelo do Veículo':^30} {'Nome do Cliente':^30}")
    print("=" * 160)    
    
    for servico in servicos:
        id_servico, descricao, preco, data_servico, data_garantia, veiculo_id, modelo_veiculo, nome_cliente = servico
        
        # Formata as datas de YYYY-MM-DD para DD/MM/YYYY
        data_servico_formatada = datetime.strptime(data_servico, '%Y-%m-%d').strftime('%d/%m/%Y')
        data_garantia_formatada = datetime.strptime(data_garantia, '%Y-%m-%d').strftime('%d/%m/%Y')
        
        print(f"{id_servico:^15} {descricao:^30} {data_servico_formatada:^15} {data_garantia_formatada:^15} R$ {preco:^10.2f} {veiculo_id:^15} {modelo_veiculo:^30} {nome_cliente:^30}")

def excluir_cliente(cliente_id):
    conn = conectar_db()
    cursor = conn.cursor()
    
    # Verifica se o cliente possui veículos associados
    cursor.execute('''
    SELECT COUNT(*) FROM veiculos WHERE cliente_id = ?
    ''', (cliente_id,))
    count = cursor.fetchone()[0]
    
    if count > 0:
        print("Erro: O cliente não pode ser excluído pois possui veículos cadastrados.")
    else:
        cursor.execute('''
        DELETE FROM clientes WHERE id = ?
        ''', (cliente_id,))
        print("Cliente excluído com sucesso.")
    
    conn.commit()
    conn.close()

def excluir_veiculo(placa):
    conn = conectar_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM veiculo WHERE placa = ?', (placa,))
    conn.commit()
    conn.close()
    print(f"Veículo com placa {placa} excluído com sucesso.")

def gerar_relatorio_servicos(data_inicio, data_fim):
    # Aqui você deve conectar ao banco de dados e buscar os serviços realizados
    conn = conectar_db()
    cursor = conn.cursor()
    
    # Exemplo de consulta para buscar serviços entre as datas especificadas
    cursor.execute('''
    SELECT s.id, s.descricao, s.preco, s.data_servico, s.data_garantia, v.id AS veiculo_id, c.nome
    FROM servicos s
    JOIN veiculos v ON s.veiculo_id = v.id
    JOIN clientes c ON v.cliente_id = c.id
    WHERE s.data_servico BETWEEN ? AND ?
    ''', (data_inicio, data_fim))
    
    servicos = cursor.fetchall()
    conn.close()

    # Exibe os resultados em formato de tabela
    print("\nRelatório de Serviços Realizados")
    print("=" * 130)
    print(f"{'ID':^5} {'Descrição':^30} {'Preço':^10} {'Data Serviço':^15} {'Data Garantia':^15} {'ID Veículo':^15} {'Nome do Cliente':^30}")
    print("=" * 130)
    
    for servico in servicos:
        id_servico, descricao, preco, data_servico, data_garantia, veiculo_id, nome_cliente = servico
        print(f"{id_servico:^5} {descricao:^30} R$ {preco:^10.2f} {data_servico:^15} {data_garantia:^15} {veiculo_id:^15} {nome_cliente:^30}")

    if not servicos:
        print("Nenhum serviço encontrado para o período informado.")

def exportar_dados():
    # Solicita ao usuário o diretório onde os arquivos serão salvos
    pasta = input("Informe o caminho da pasta onde deseja salvar os arquivos: ")
    
    # Verifica se a pasta existe
    if not os.path.exists(pasta):
        print("A pasta informada não existe. Por favor, verifique o caminho.")
        return

    # Exportar Clientes
    exportar_clientes(pasta)
    # Exportar Veículos
    exportar_veiculos(pasta)
    # Exportar Serviços
    exportar_servicos(pasta)

def exportar_clientes(pasta):
    conn = conectar_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM clientes')
    clientes = cursor.fetchall()
    conn.close()

    # Cria um novo arquivo XLSX para clientes
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Adiciona cabeçalho
    ws.append(['ID', 'Nome', 'CPF', 'Endereço'])

    # Adiciona os dados dos clientes
    for cliente in clientes:
        ws.append(cliente)

    # Salva o arquivo
    wb.save(os.path.join(pasta, 'Clientes.xlsx'))  # Alterado para .xlsx
    print("Arquivo de Clientes exportado com sucesso!")

def exportar_veiculos(pasta):
    conn = conectar_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM veiculos')
    veiculos = cursor.fetchall()
    conn.close()

    # Cria um novo arquivo XLSX para veículos
    wb = Workbook()
    ws = wb.active
    ws.title = "Veículos"

    # Adiciona cabeçalho
    ws.append(['ID', 'Cliente ID', 'Modelo', 'Placa', 'Fabricante', 'Ano'])

    # Adiciona os dados dos veículos
    for veiculo in veiculos:
        ws.append(veiculo)

    # Salva o arquivo
    wb.save(os.path.join(pasta, 'Veiculos.xlsx'))  # Alterado para .xlsx
    print("Arquivo de Veículos exportado com sucesso!")

def exportar_servicos(pasta):
    conn = conectar_db()
    cursor = conn.cursor()
    
    # Faz uma junção entre as tabelas servicos, veiculos e clientes
    cursor.execute('''
    SELECT s.id, s.descricao, s.preco, s.data_servico, s.data_garantia, v.id AS veiculo_id, v.modelo, c.nome
    FROM servicos s
    JOIN veiculos v ON s.veiculo_id = v.id
    JOIN clientes c ON v.cliente_id = c.id
    ''')
    
    servicos = cursor.fetchall()
    conn.close()

    # Cria um novo arquivo XLSX para serviços
    wb = Workbook()
    ws = wb.active
    ws.title = "Serviços"

    # Adiciona cabeçalho
    ws.append(['ID', 'Descrição', 'Preço', 'Data Serviço', 'Data Garantia', 'ID Veículo', 'Modelo do Veículo', 'Nome do Cliente'])

    # Adiciona os dados dos serviços
    for servico in servicos:
        ws.append(servico)

    # Salva o arquivo
    wb.save(os.path.join(pasta, 'Servicos.xlsx'))  # Alterado para .xlsx
    print("Arquivo de Serviços exportado com sucesso!")

def main():
    criar_tabelas()  # Certifique-se de que as tabelas estão criadas

    while True:
        try:
            opcao = menu()

            if opcao == "1":
                cadastrar_cliente()
            elif opcao == "2":
                cadastrar_veiculo()
            elif opcao == "3":
                registrar_servico()
            elif opcao == "4":
                listar_clientes()
            elif opcao == "5":
                listar_veiculos()
            elif opcao == "6":
                listar_servicos()
            elif opcao == "7":
                cliente_id = input("Informe o ID do cliente a ser excluído: ")
                excluir_cliente(cliente_id)
            elif opcao == "8":
                placa = input("Informe a placa do veículo a ser excluído: ")
                excluir_veiculo(placa)
            elif opcao == "9":
                exportar_ou_gerar_relatorio()  # Chama a nova função
            elif opcao == "q":
                print("Saindo do sistema...")
                break
            else:
                print("Opção inválida. Tente novamente.")
        
        except Exception as e:
            print(f"Ocorreu um erro: {e}. Retornando ao menu principal.")

def exportar_ou_gerar_relatorio():
    print("\nEscolha uma opção:")
    print("1. Gerar Relatório de Serviços Realizados")
    print("2. Exportar Dados (Arquivos Excel)")
    
    escolha = input("Escolha uma opção: ")
    
    if escolha == '1':
        data_inicio = input("Informe a data de início (DD/MM/YYYY): ")
        data_fim = input("Informe a data de fim (DD/MM/YYYY): ")
        gerar_relatorio_servicos(data_inicio, data_fim)  # Chama a função para gerar o relatório
    elif escolha == '2':
        exportar_dados()  # Chama a função de exportação
    else:
        print("Opção inválida. Retornando ao menu principal.")

if __name__ == "__main__":
    main()